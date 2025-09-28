import os
import math
import json
import subprocess
import ezdxf
import pandas as pd
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import Workbook

# ---------------- Load config ----------------
with open("config.json", "r") as f:
    config = json.load(f)

INPUT_FOLDER = config.get("input_folder", r"C:\PIDs")
ODA_CONVERTER = config.get("oda_converter", r"C:\Program Files/ODA/ODAFileConverter.exe")
NEARBY_TEXT_RADIUS = config.get("nearby_text_radius", 200)
DXF_VERSION = config.get("dxf_version", "ACAD2013")

OUTPUT_FILE = os.path.join(INPUT_FOLDER, "MTO.xlsx")
LOG_FILE = os.path.join(INPUT_FOLDER, "MTO_log.txt")

mto_data = []
file_status = []
log_lines = []


def log(msg):
    print(msg)
    log_lines.append(msg)


def distance(p1, p2):
    return math.dist(p1, p2)


def find_nearby_text(msp, insert_point, radius=200):
    texts = []
    for txt in msp.query("TEXT MTEXT"):
        try:
            pos = txt.dxf.insert if txt.dxftype() == "TEXT" else txt.dxf.insert
            if distance(insert_point, pos) <= radius:
                texts.append(txt.plain_text() if txt.dxftype() == "MTEXT" else txt.dxf.text)
        except Exception:
            continue
    return texts


def process_dxf(file_path, source_type="DXF", converted=False):
    try:
        doc = ezdxf.readfile(file_path)
    except Exception as e:
        log(f"❌ Could not read {file_path}: {e}")
        file_status.append({
            "File": os.path.basename(file_path),
            "Type": source_type,
            "Status": "Failed",
            "BlocksExtracted": 0
        })
        return

    msp = doc.modelspace()
    count = 0

    for entity in msp.query("INSERT"):
        block_name = entity.dxf.name.upper()
        insert_point = entity.dxf.insert

        attribs = {}
        if entity.has_attribs():
            attribs = {att.dxf.tag.upper(): att.dxf.text for att in entity.attribs()}

        size = attribs.get("SIZE", "")
        rating = attribs.get("RATING", "")
        spec = attribs.get("SPEC", "")
        tag = attribs.get("TAG", "")

        if not (size and rating and spec):
            nearby_texts = find_nearby_text(msp, insert_point, NEARBY_TEXT_RADIUS)
            for txt in nearby_texts:
                t = txt.strip().upper()
                if "DN" in t or '"' in t:
                    size = size or t
                elif "#" in t or "CLASS" in t:
                    rating = rating or t
                elif "CS" in t or "SS" in t or "A" in t:
                    spec = spec or t

        if not (size and rating and spec):
            log(f"⚠️ Missing attributes for {block_name} in {os.path.basename(file_path)}")

        mto_data.append({
            "Drawing": os.path.basename(file_path),
            "BlockName": block_name,
            "SIZE": size,
            "RATING": rating,
            "SPEC": spec,
            "TAG": tag
        })
        count += 1

    status = "Processed"
    if converted:
        status = "Converted+Processed"

    log(f"✅ Processed {file_path} ({count} blocks)")
    file_status.append({
        "File": os.path.basename(file_path),
        "Type": source_type,
        "Status": status,
        "BlocksExtracted": count
    })


def convert_dwg_to_dxf(dwg_path):
    dxf_path = dwg_path.replace(".dwg", ".dxf")
    try:
        subprocess.run(
            [ODA_CONVERTER, dwg_path, os.path.dirname(dxf_path), DXF_VERSION, "DXF", "1"],
            check=True
        )
        log(f"🔄 Converted {dwg_path} → {dxf_path} ({DXF_VERSION})")
        return dxf_path
    except Exception as e:
        log(f"❌ Failed to convert {dwg_path}: {e}")
        file_status.append({
            "File": os.path.basename(dwg_path),
            "Type": "DWG",
            "Status": "Conversion Failed",
            "BlocksExtracted": 0
        })
        return None


def scan_folder(path):
    for root, _, files in os.walk(path):
        for fname in files:
            fpath = os.path.join(root, fname)
            if fname.lower().endswith(".dxf"):
                process_dxf(fpath, source_type="DXF")
            elif fname.lower().endswith(".dwg"):
                dxf = convert_dwg_to_dxf(fpath)
                if dxf:
                    process_dxf(dxf, source_type="DWG", converted=True)


# ---------------- Run Script ----------------
log(f"\n=== MTO Extraction Started: {datetime.now()} ===\n")
scan_folder(INPUT_FOLDER)

# ---------------- Export Excel with formatting ----------------
if mto_data:
    df_detailed = pd.DataFrame(mto_data)
    df_grouped = df_detailed.groupby(["BlockName", "SIZE", "RATING", "SPEC"]).size().reset_index(name="Count")
    df_grouped = df_grouped.sort_values(["BlockName", "SIZE"])
    df_status = pd.DataFrame(file_status)

    wb = Workbook()
    wb.remove(wb.active)

    # Detailed_Data sheet
    ws1 = wb.create_sheet("Detailed_Data")
    for r in dataframe_to_rows(df_detailed, index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    ws1.auto_filter.ref = ws1.dimensions

    # Grouped_MTO sheet
    ws2 = wb.create_sheet("Grouped_MTO")
    for r in dataframe_to_rows(df_grouped, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.auto_filter.ref = ws2.dimensions

    # File_Status sheet
    ws3 = wb.create_sheet("File_Status")
    for r in dataframe_to_rows(df_status, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.auto_filter.ref = ws3.dimensions

    wb.save(OUTPUT_FILE)
    log(f"\n📊 MTO exported to {OUTPUT_FILE}")
else:
    log("\n⚠️ No MTO data extracted.")

# Save log file
with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))
log(f"📝 Log saved to {LOG_FILE}")