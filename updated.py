import os
import json
import math
import subprocess
import ezdxf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ----------------- CONFIG -----------------
CONFIG_FILE = "config.json"
DEFAULT_CONFIG = {
    "input_folder": "C:/PIDs",
    "oda_converter": "C:/Program Files/ODA/ODAFileConverter.exe",
    "output_excel": "MTO.xlsx",
    "output_log": "MTO_log.txt",
    "dxf_version": "ACAD2013",
    "nearby_text_radius": 200.0
}

# ----------------- GLOBALS -----------------
mto_data = []
status_data = []
NEARBY_TEXT_RADIUS = DEFAULT_CONFIG["nearby_text_radius"]

# ----------------- HELPERS -----------------
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            config = json.load(f)
    else:
        config = DEFAULT_CONFIG
        with open(CONFIG_FILE, "w") as f:
            json.dump(config, f, indent=4)
    global NEARBY_TEXT_RADIUS
    NEARBY_TEXT_RADIUS = config.get("nearby_text_radius", 200.0)
    return config

def log(message):
    print(message)
    with open(config["output_log"], "a") as f:
        f.write(f"{datetime.now()} - {message}\n")

def distance(p1, p2):
    return math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2)

def find_nearby_text(msp, point, radius=200):
    texts = []
    for e in msp.query("TEXT MTEXT"):
        try:
            txt_point = e.dxf.insert if e.dxftype() == "TEXT" else e.dxf.insert
            d = distance(point, txt_point)
            if d <= radius:
                texts.append(e.text if e.dxftype() == "TEXT" else e.plain_text())
        except:
            pass
    return texts

# ----------------- PIPE HANDLER -----------------
def process_pipes(msp, drawing_name):
    count = 0
    total_length = 0

    for ent in msp.query("LINE LWPOLYLINE"):
        length = 0
        if ent.dxftype() == "LINE":
            p1, p2 = ent.dxf.start, ent.dxf.end
            length = distance(p1, p2)
            insert_point = p1
        elif ent.dxftype() == "LWPOLYLINE":
            points = list(ent.get_points("xy"))
            for i in range(len(points) - 1):
                length += distance(points[i], points[i + 1])
            insert_point = points[0]
        else:
            continue

        # Nearby text for size/spec
        size, spec = "", ""
        nearby_texts = find_nearby_text(msp, insert_point, NEARBY_TEXT_RADIUS)
        for txt in nearby_texts:
            t = txt.strip().upper()
            if "DN" in t or '"' in t:
                size = size or t
            elif "CS" in t or "SS" in t or "A" in t:
                spec = spec or t

        mto_data.append({
            "Drawing": drawing_name,
            "Type": "Pipe",
            "BlockName": "",
            "SIZE": size,
            "RATING": "",
            "SPEC": spec,
            "TAG": "",
            "LENGTH": round(length, 2)
        })
        count += 1
        total_length += length

    return count, total_length

# ----------------- BLOCK HANDLER -----------------
def process_blocks(msp, drawing_name):
    count = 0
    for entity in msp.query("INSERT"):
        block_name = entity.dxf.name
        size, rating, spec, tag = "", "", "", ""
        if entity.has_attribs():
            for attrib in entity.attribs:
                tagname = attrib.dxf.tag.upper()
                val = attrib.dxf.text
                if "SIZE" in tagname:
                    size = val
                elif "RATING" in tagname:
                    rating = val
                elif "SPEC" in tagname:
                    spec = val
                elif "TAG" in tagname:
                    tag = val

        mto_data.append({
            "Drawing": drawing_name,
            "Type": "Block",
            "BlockName": block_name,
            "SIZE": size,
            "RATING": rating,
            "SPEC": spec,
            "TAG": tag,
            "LENGTH": ""
        })
        count += 1
    return count

# ----------------- DXF HANDLER -----------------
def process_dxf(file_path, source_type="DXF", converted=False):
    try:
        doc = ezdxf.readfile(file_path)
    except Exception as e:
        log(f"❌ Could not read {file_path}: {e}")
        return

    msp = doc.modelspace()
    drawing_name = os.path.basename(file_path)

    block_count = process_blocks(msp, drawing_name)
    pipe_count, pipe_length = process_pipes(msp, drawing_name)

    log(f"✅ {drawing_name}: {block_count} blocks, {pipe_count} pipes, {round(pipe_length,2)} units length")

    status_data.append({
        "Drawing": drawing_name,
        "Source": source_type,
        "Converted": converted,
        "BlockCount": block_count,
        "PipeCount": pipe_count,
        "PipeLength": round(pipe_length, 2)
    })

# ----------------- DWG HANDLER -----------------
def convert_dwg_to_dxf(dwg_path, dxf_path):
    cmd = [
        config["oda_converter"],
        dwg_path,
        dxf_path,
        config["dxf_version"],
        "1"
    ]
    try:
        subprocess.run(cmd, check=True)
        return True
    except Exception as e:
        log(f"❌ Conversion failed for {dwg_path}: {e}")
        return False

# ----------------- EXPORT -----------------
def export_to_excel(output_path):
    df = pd.DataFrame(mto_data)
    status_df = pd.DataFrame(status_data)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not df.empty:
            df.to_excel(writer, sheet_name="Detailed_Data", index=False)
            grouped = df.groupby(["Type","BlockName","SIZE","RATING","SPEC"], dropna=False).agg(
                Count=("BlockName","count"),
                TotalLength=("LENGTH","sum")
            ).reset_index()
            grouped.to_excel(writer, sheet_name="Grouped_MTO", index=False)

        if not status_df.empty:
            status_df.to_excel(writer, sheet_name="File_Status", index=False)

    wb = load_workbook(output_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
    wb.save(output_path)

# ----------------- MAIN -----------------
if __name__ == "__main__":
    config = load_config()
    folder = config["input_folder"]
    config["output_excel"] = os.path.join(folder, "MTO.xlsx")
    config["output_log"] = os.path.join(folder, "MTO_log.txt")

    open(config["output_log"], "w").close()  # clear log

    log("🚀 Starting MTO extraction...")

    for root, _, files in os.walk(folder):
        for file in files:
            path = os.path.join(root, file)
            ext = file.lower().split(".")[-1]
            if ext == "dxf":
                process_dxf(path, "DXF")
            elif ext == "dwg":
                dxf_path = os.path.splitext(path)[0] + ".dxf"
                if convert_dwg_to_dxf(path, dxf_path):
                    process_dxf(dxf_path, "DWG", converted=True)

    export_to_excel(config["output_excel"])
    log(f"📊 Extraction finished. Excel saved at {config['output_excel']}")